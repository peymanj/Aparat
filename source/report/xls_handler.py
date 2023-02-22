import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles.colors import RgbColor
import traceback

rgb_border = 'bdc0bf'
thin_border = Border(
                    left=Side(border_style=BORDER_THIN, color=openpyxl.styles.colors.Color(rgb=rgb_border)),
                    right=Side(border_style=BORDER_THIN, color=openpyxl.styles.colors.Color(rgb=rgb_border)),
                    top=Side(border_style=BORDER_THIN, color=openpyxl.styles.colors.Color(rgb=rgb_border)),
                    bottom=Side(border_style=BORDER_THIN, color=openpyxl.styles.colors.Color(rgb=rgb_border))
                )


def xls_write(result, loc): # mode= append or integrate
 
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "New Title"

    for row in result:
        sheet.append(row)

    styler(sheet)
    wb.save(loc)



# def styler(sheet, symbol_indexes):
def styler(sheet):
    try:

        first_col = 1
        last_col = sheet.max_column + 1

        mr = sheet.max_row
        for col in range(first_col, last_col):
            for row in range(1, mr + 1):

                sheet.cell(row=row, column=col).border = thin_border

                if row % 2 == 0:
                    bg_color = openpyxl.styles.colors.Color(rgb='bdc0bf')
                else:
                    bg_color = openpyxl.styles.colors.Color(rgb='f6f6f6')
                    

                fill_string = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=bg_color)
                sheet.cell(row=row, column=col).fill = fill_string


    except Exception as e:
        traceback.print_exc()
        print(e)
        print(f"Error occurred while applying styles to file")
        raise e
